import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

os.makedirs('sampleFile', exist_ok=True)

# ── FILE 1: messy_invoices.xlsx ──────────────────────────────────────────────
wb1 = Workbook()
ws = wb1.active
ws.title = 'Q1 Invoices'
ws['A1'] = 'Q1 Invoice Report - Acme Corp'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:I1')
ws.append([])
headers = ['Inv#', 'Client Name', 'Issue Date', 'Due Date', 'Amount', 'Tax', 'Total', 'Status', 'Notes']
ws.append(headers)
for cell in ws[3]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color='4472C4')

rows = [
    ['INV-001', 'Northwind Traders', '01/05/2025', '01/20/2025', '$5,200.00', '$520.00', '$5,720.00', 'Paid', 'Early payment discount applied'],
    ['INV-002', 'Contoso Ltd', 'January 8 2025', '2025-01-23', 'USD 3400', '340', 'USD 3740', 'PAID', ''],
    ['INV-003', 'Fabrikam Inc.', '2025-01-10', '25/01/2025', '8750', '875.00', '9625', 'Unpaid', 'Net 15 terms'],
    ['INV-004', 'Adventure Works', '12 Jan 2025', '01/27/2025', '$12,000.00', '$1,200.00', '$13,200.00', 'paid', 'Recurring monthly'],
    ['INV-005', 'Tailspin Toys', '15/01/2025', 'January 30 2025', '$2,150', '$215', '$2365', 'Pending', 'Awaiting PO confirmation'],
    ['INV-006', 'Wide World Importers', '2025-01-18', '02/02/2025', 'USD 6600', '660.00', '$7,260', 'UNPAID', ''],
    [],
    ['INV-007', 'Southridge Video', '20 January 2025', '2025-02-04', '$4,400.00', '$440', '4840', 'pnding', 'Check with AM'],
    ['INV-008', 'Blue Yonder Airlines', '01/22/2025', '2025-02-06', '$18,500.00', '$1,850.00', '$20,350.00', 'Paid', 'Wire transfer confirmed'],
    ['INV-009', 'Graphic Design Inst.', 'January 25 2025', '02/09/2025', '1200', '120', '1320', 'Unpaid', ''],
    ['INV-010', 'Lucerne Publishing', '2025-01-28', '12 Feb 2025', '$9,750.00', '975.00', 'USD 10725', 'paid', 'Invoice resent x2'],
    ['INV-011', 'Coho Winery', '30/01/2025', '2025-02-14', '$3,800', '$380.00', '$4,180', 'PAID', ''],
    ['INV-012', 'Humongous Insurance', '2025-02-01', 'Feb 16 2025', 'USD 22000', '2200', '24200', 'Pending', 'Legal review required'],
    [],
    ['INV-013', 'City Power & Light', '03 Feb 2025', '2025-02-18', '$7,100.00', '$710', '$7,810.00', 'Unpaid', 'Dispute raised'],
    ['INV-014', 'Datum Corporation', '2025-02-05', '02/20/2025', '$5,500.00', '$550.00', '$6,050.00', 'paid', ''],
    ['INV-015', 'Margie Travel', 'February 7 2025', '22/02/2025', '$1,875', '187.50', '$2,062.50', 'PAID', 'Travel reimbursement'],
    ['INV-016', 'Proseware Inc', '2025-02-10', '2025-02-25', 'USD 14300', '1430.00', 'USD 15730', 'Unpaid', ''],
    ['INV-017', 'School of Fine Art', '12 February 2025', 'Feb 27 2025', '$4,200.00', '$420', '4620', 'pnding', 'Waiting on signed contract'],
    ['INV-018', 'Wingtip Toys', '2025-02-14', '01/03/2025', '$6,600.00', '$660.00', '$7,260.00', 'UNPAID', ''],
    ['INV-019', 'Trey Research', '16/02/2025', '2025-03-03', '$3,150', '315', 'USD 3465', 'Paid', 'Overpaid - credit note issued'],
    [],
    ['INV-020', 'Bellows College', '2025-02-18', 'March 5 2025', '$11,000.00', '$1,100.00', '$12,100.00', 'Pending', ''],
    ['INV-021', 'A Datum Corporation', 'February 20 2025', '2025-03-07', '$2,700', '$270.00', '$2,970', 'paid', 'Net 15'],
    ['INV-022', 'Consolidated Msg', '22/02/2025', '09/03/2025', 'USD 8900', '890', 'USD 9790', 'PAID', 'Direct debit'],
    ['INV-023', 'Fourth Coffee', '2025-02-24', '11 March 2025', '$16,500.00', '$1,650.00', '$18,150.00', 'Unpaid', ''],
    ['INV-024', 'Relecloud', 'Feb 26 2025', '2025-03-13', '$5,250', '525.00', '$5,775', 'paid', ''],
    ['INV-025', 'Northwind Traders', '2025-02-28', '2025-03-15', '$5,400.00', '$540', '$5,940.00', 'UNPAID', 'Follow-up sent 3/1'],
]
for r in rows:
    ws.append(r)
for col, w in zip('ABCDEFGHI', [12, 25, 18, 18, 14, 12, 14, 12, 35]):
    ws.column_dimensions[col].width = w
wb1.save('sampleFile/messy_invoices.xlsx')
print('OK messy_invoices.xlsx')

# ── FILE 2: employee_records.xlsx ────────────────────────────────────────────
wb2 = Workbook()
ws1 = wb2.active
ws1.title = 'Active Employees'
ws1.append(['Emp_ID', 'Full_Name', 'Department', 'Hire_Date', 'Salary', 'Email', 'Phone', 'Manager_ID', 'Location'])
for c in ws1[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='217346')

emp_rows = [
    ['E001', 'James Harrington', 'Engineering', '2019-03-15', 95000, 'j.harrington@acme.com', '555-0101', 'E010', 'New York'],
    ['E002', 'Priya Sharma', 'Engineering', '2020-07-22', 88000, 'p.sharma@acme.com', '555-0102', 'E010', 'Remote'],
    ['E003', 'Marcus Webb', 'Sales', '2018-11-01', 72000, 'm.webb@acme.com', '555-0103', 'E011', 'Chicago'],
    ['E004', 'Sofia Delgado', 'Marketing', '2021-01-10', 68000, 's.delgado@acme.com', '555-0104', 'E012', 'Los Angeles'],
    ['E005', 'Chen Wei', 'Engineering', '2017-06-30', 102000, 'c.wei@acme.com', '555-0105', 'E010', 'New York'],
    ['E006', 'Amara Okafor', 'HR', '2022-03-14', 61000, 'a.okafor@acme.com', '555-0106', 'E013', 'Chicago'],
    ['E007', 'Ryan OBrien', 'Sales', '2020-09-05', 75000, 'r.obrien@acme.com', '555-0107', 'E011', 'Remote'],
    ['E008', 'Yuki Tanaka', 'Finance', '2019-12-01', 84000, 'y.tanaka@acme.com', '555-0108', 'E014', 'New York'],
    ['E009', 'Fatima Al-Rashid', 'Marketing', '2023-02-20', 65000, 'f.alrashid@acme.com', '555-0109', 'E012', 'Los Angeles'],
    ['E010', 'Daniel Kovacs', 'Engineering', '2015-04-15', 145000, 'd.kovacs@acme.com', '555-0110', 'E020', 'New York'],
    ['E011', 'Sarah Blackwell', 'Sales', '2016-08-22', 138000, 's.blackwell@acme.com', '555-0111', 'E020', 'Chicago'],
    ['E012', 'Liam Nguyen', 'Marketing', '2017-02-14', 128000, 'l.nguyen@acme.com', '555-0112', 'E020', 'Los Angeles'],
    ['E013', 'Aisha Patel', 'HR', '2018-05-30', 118000, 'a.patel@acme.com', '555-0113', 'E020', 'Chicago'],
    ['E014', 'Robert Kim', 'Finance', '2016-10-12', 132000, 'r.kim@acme.com', '555-0114', 'E020', 'New York'],
    ['E015', 'Emma Johansson', 'Engineering', '2021-08-16', 86000, 'e.johansson@acme.com', '555-0115', 'E010', 'Remote'],
    ['E016', 'Carlos Mendez', 'Sales', '2022-05-03', 71000, 'c.mendez@acme.com', '555-0116', 'E011', 'Chicago'],
    ['E017', 'Nina Petrov', 'Finance', '2020-11-23', 79000, 'n.petrov@acme.com', '555-0117', 'E014', 'New York'],
    ['E018', 'David Osei', 'Engineering', '2023-06-01', 82000, 'd.osei@acme.com', '555-0118', 'E010', 'New York'],
    ['E019', 'Hannah Schmidt', 'Marketing', '2019-04-18', 70000, 'h.schmidt@acme.com', '555-0119', 'E012', 'Remote'],
    ['E020', 'Victoria Cross', 'Executive', '2012-01-09', 210000, 'v.cross@acme.com', '555-0120', None, 'New York'],
    ['E021', 'Omar Hassan', 'Engineering', '2022-09-12', 87000, 'o.hassan@acme.com', '555-0121', 'E010', 'Remote'],
    ['E022', 'Beatrice Laurent', 'Sales', '2021-03-25', 74000, 'b.laurent@acme.com', '555-0122', 'E011', 'Los Angeles'],
    ['E023', 'Kwame Asante', 'Finance', '2020-07-08', 81000, 'k.asante@acme.com', '555-0123', 'E014', 'Chicago'],
    ['E024', 'Ingrid Lindqvist', 'HR', '2023-01-16', 62000, 'i.lindqvist@acme.com', '555-0124', 'E013', 'Remote'],
    ['E025', 'Patrick Flanagan', 'Marketing', '2018-10-29', 73000, 'p.flanagan@acme.com', '555-0125', 'E012', 'Los Angeles'],
    ['E026', 'Mei Chen', 'Engineering', '2022-12-05', 84000, 'm.chen@acme.com', '555-0126', 'E010', 'New York'],
    ['E027', 'Isaac Abramowitz', 'Sales', '2019-02-14', 77000, 'i.abramowitz@acme.com', '555-0127', 'E011', 'Chicago'],
    ['E028', 'Alicia Torres', 'Finance', '2021-06-28', 76000, 'a.torres@acme.com', '555-0128', 'E014', 'New York'],
    ['E029', 'Benjamin Fox', 'HR', '2020-04-07', 63000, 'b.fox@acme.com', '555-0129', 'E013', 'Chicago'],
    ['E030', 'Nadia Volkov', 'Engineering', '2017-11-19', 98000, 'n.volkov@acme.com', '555-0130', 'E010', 'Remote'],
]
for r in emp_rows:
    ws1.append(r)
for col, w in zip(range(1, 10), [8, 22, 16, 14, 12, 30, 14, 12, 14]):
    ws1.column_dimensions[get_column_letter(col)].width = w

ws2 = wb2.create_sheet('Terminated')
ws2.append(['employee_number', 'name', 'dept', 'termination_date', 'final_salary', 'contact_email'])
for c in ws2[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='C00000')
term_rows = [
    ['T001', 'Greg Holloway', 'Sales', '2024-03-31', 68000, 'g.holloway@personal.com'],
    ['T002', 'Diane Marsh', 'Engineering', '2024-01-15', 91000, 'dmarsh@gmail.com'],
    ['T003', 'Frank Russo', 'HR', '2023-11-30', 59000, 'f.russo@yahoo.com'],
    ['T004', 'Sandra Lee', 'Marketing', '2024-06-14', 66000, 'sandra.lee@outlook.com'],
    ['T005', 'Paul Theron', 'Finance', '2023-09-22', 78000, 'p.theron@personal.com'],
    ['T006', 'Janet Mills', 'Engineering', '2024-02-29', 89000, 'j.mills@gmail.com'],
    ['T007', 'Kevin Park', 'Sales', '2024-04-12', 71000, 'kpark@hotmail.com'],
    ['T008', 'Laura Benson', 'Finance', '2023-12-01', 83000, 'l.benson@gmail.com'],
    ['T009', 'Tom Graves', 'Engineering', '2024-05-31', 93000, 'tgraves@personal.com'],
    ['T010', 'Olivia Stern', 'HR', '2024-01-31', 60000, 'ostern@yahoo.com'],
    ['T011', 'Henry Drake', 'Marketing', '2023-10-15', 67000, 'h.drake@gmail.com'],
    ['T012', 'Cynthia Hall', 'Sales', '2024-03-15', 72000, 'c.hall@outlook.com'],
    ['T013', 'Albert Ng', 'Finance', '2024-06-30', 80000, 'a.ng@personal.com'],
    ['T014', 'Rosa Gomez', 'Engineering', '2023-08-31', 88000, 'r.gomez@gmail.com'],
    ['T015', 'William Stone', 'Executive', '2024-06-01', 195000, 'w.stone@personal.com'],
]
for r in term_rows:
    ws2.append(r)
for col, w in zip(range(1, 7), [16, 22, 16, 18, 14, 28]):
    ws2.column_dimensions[get_column_letter(col)].width = w

ws3 = wb2.create_sheet('Summary')
ws3['A1'] = 'HEADCOUNT & SALARY SUMMARY'
ws3['A1'].font = Font(bold=True, size=13)
ws3.merge_cells('A1:E1')
ws3.append([])
ws3.append(['Department', 'Active', 'Terminated', 'Avg Active Salary', 'Avg Term Salary'])
for c in ws3[3]:
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', start_color='DDEBF7')
for r in [
    ['Engineering', 11, 4, 91375, 90250],
    ['Sales', 5, 3, 81600, 70333],
    ['Marketing', 4, 2, 76500, 66500],
    ['Finance', 4, 3, 90500, 80333],
    ['HR', 3, 2, 80667, 59500],
    ['Executive', 1, 1, 210000, 195000],
]:
    ws3.append(r)
for col, w in zip(range(1, 6), [18, 10, 14, 22, 18]):
    ws3.column_dimensions[get_column_letter(col)].width = w
wb2.save('sampleFile/employee_records.xlsx')
print('OK employee_records.xlsx')

# ── FILE 3: sales_pipeline.xlsx ──────────────────────────────────────────────
wb3 = Workbook()
ws = wb3.active
ws.title = 'Pipeline'
ws.append(['Deal ID', 'Company', 'Contact Person', 'Contact Email', 'Deal Value', 'Stage',
           'Probability%', 'Close Date', 'Sales Rep', 'Region', 'Product', 'Notes'])
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='7030A0')

pipeline = [
    ['D-001','Globex Corp','Homer Simpson','h.simpson@globex.com','$450,000','Closed Won','100%','2025-01-15','Sarah Blackwell','North East','Enterprise Suite','Upsell Q2'],
    ['D-002','Initech','Bill Lumbergh','b.lumbergh@initech.com','USD 120000','Proposal','75%','March 31 2025','Marcus Webb','Mid West','Professional','TPS reports integration'],
    ['D-003','Umbrella Corp','Albert Wesker','a.wesker@umbrella.com','$2,200,000','Negotiation','0.85','15/04/2025','Sarah Blackwell','South East','Enterprise Suite','Legal review in progress'],
    ['D-004','Soylent Corp','Jonathan Salk','j.salk@soylent.com','320000','Discovery','medium','30 Apr 2025','Ryan OBrien','West','Starter','Price sensitivity flagged'],
    ['D-005','Gekko & Co','Gordon Gekko','g.gekko@gekkoco.com','$875,000','Closed Won','100%','2025-02-28','Carlos Mendez','North East','Professional','Multi-year 3yr'],
    ['D-006','Vandelay Industries','Art Vandelay','a.vandelay@vandelay.com',None,'Prospecting','low',None,'Marcus Webb','Mid West','Starter','Cold outbound'],
    ['D-007','Dunder Mifflin','Michael Scott','m.scott@dundermifflin.com','USD 95000','Proposal','60%','2025-05-15','Ryan OBrien','North East','Professional','Paper supply expansion'],
    ['D-008','Sterling Cooper','Don Draper','d.draper@sterlingcooper.com','$1,500,000','Negotiation','0.80','May 30 2025','Sarah Blackwell','North East','Enterprise Suite','Creative agency'],
    ['D-009','Pied Piper','Richard Hendricks','r.hendricks@piedpiper.com','185,000','Discovery','25%','15 Jun 2025','Beatrice Laurent','West','Professional','Startup pricing'],
    ['D-010','Hooli','Gavin Belson','g.belson@hooli.com','$5,000,000','Proposal','high','2025-06-30','Sarah Blackwell','West','Enterprise Suite','Competitive vs Nucleus'],
    ['D-011','Nakatomi Corp','John McClane','j.mcclane@nakatomi.com','USD 230000','Closed Won','100%','2025-03-01','Carlos Mendez','West','Professional',''],
    ['D-012','Prestige Worldwide','Step Brothers','info@prestigeww.com',None,'Prospecting','5%',None,'Marcus Webb','South East','Starter','Unclear budget'],
    ['D-013','Cyberdyne Systems','Miles Dyson','m.dyson@cyberdyne.com','$3,800,000','Negotiation','0.70','31/07/2025','Sarah Blackwell','West','Enterprise Suite','AI/automation expansion'],
    ['D-014','Wernham Hogg','David Brent','d.brent@wernhamhogg.com','72000','Proposal','50%','August 15 2025','Ryan OBrien','International','Starter','UK market pilot'],
    ['D-015','Tyrell Corp','Eldon Tyrell','e.tyrell@tyrell.com','$4,250,000','Discovery','medium','2025-09-30','Sarah Blackwell','West','Enterprise Suite','Replicant workforce mgmt'],
    ['D-016','Buy More','Chuck Bartowski','c.bartowski@buymore.com','USD 55000','Closed Lost','0%','2025-01-20','Carlos Mendez','West','Starter','Lost to competitor'],
    ['D-017','Paper Street Soap','Tyler Durden','t.durden@paperstreet.com','$180,000','Proposal','65%','30 Sep 2025','Beatrice Laurent','North East','Professional','Unusual requirements'],
    ['D-018','Black Mesa','Gordon Freeman','g.freeman@blackmesa.com','$2,900,000','Negotiation','0.75','2025-10-15','Sarah Blackwell','South East','Enterprise Suite','Research division'],
    ['D-019','Los Pollos Hermanos','G Fring','g.fring@lph.com','USD 410000','Discovery','30%','October 31 2025','Marcus Webb','South East','Professional','Multi-location rollout'],
    ['D-020','Monsters Inc','James Sullivan','j.sullivan@monstersinc.com','$620,000','Proposal','70%','15/11/2025','Ryan OBrien','West','Professional','Energy sector pivot'],
    ['D-021','Initech','Peter Gibbons','p.gibbons@initech.com','90000','Closed Lost','0%','2025-02-14','Marcus Webb','Mid West','Starter','Internal champion left'],
    ['D-022','Weyland-Yutani','Carter Burke','c.burke@weyland.com','$6,100,000','Discovery','high','2025-11-30','Sarah Blackwell','International','Enterprise Suite','Space exploration'],
    ['D-023','Wonka Industries','Willy Wonka','w.wonka@wonka.com','USD 340000','Proposal','55%','December 15 2025','Beatrice Laurent','North East','Professional','Unusual procurement'],
    ['D-024','Virtucon','Scott Evil','s.evil@virtucon.com','$1,100,000','Negotiation','0.65','31 Dec 2025','Sarah Blackwell','International','Enterprise Suite',''],
    ['D-025','Springfield Nuclear','Montgomery Burns','m.burns@snpp.com',None,'Prospecting','low',None,'Carlos Mendez','Mid West','Enterprise Suite','Needs board approval'],
    ['D-026','Acme Corporation','Wile E Coyote','w.coyote@acme.com','$275,000','Closed Won','100%','2025-03-22','Marcus Webb','West','Professional',''],
    ['D-027','Gringotts Bank','Griphook','g.riphook@gringotts.com','USD 8500000','Discovery','medium','2026-01-31','Sarah Blackwell','International','Enterprise Suite','Financial services'],
    ['D-028','Sirius Cybernetics','Marvin Android','m.android@sirius.com','$490,000','Proposal','45%','2026-02-28','Ryan OBrien','International','Professional','Robot integration'],
    ['D-029','Duff Beer Co','Duff Man','d.man@duffbeer.com','150000','Discovery','20%','February 28 2026','Beatrice Laurent','Mid West','Starter','POC required first'],
    ['D-030','Rekall Inc','Douglas Quaid','d.quaid@rekall.com','$720,000','Negotiation','0.80','2026-03-15','Sarah Blackwell','West','Professional','Memory implant workflow'],
    ['D-031','MomCorp','Mom','mom@momcorp.com','$12,000,000','Discovery','high',None,'Sarah Blackwell','International','Enterprise Suite','World domination tier'],
    ['D-032','Planet Express','Philip Fry','p.fry@planetexpress.com','USD 210000','Proposal','40%','31/03/2026','Carlos Mendez','International','Professional','Intergalactic routes'],
    ['D-033','Hanso Foundation','Thomas Mittelwerk','t.mittelwerk@hanso.com','$3,300,000','Negotiation','0.60','2026-04-15','Sarah Blackwell','International','Enterprise Suite','Island deployment'],
    ['D-034','Stark Industries','Tony Stark','t.stark@stark.com','$25,000,000','Discovery','medium',None,'Sarah Blackwell','North East','Enterprise Suite','Iron Man fleet mgmt'],
    ['D-035','Wayne Enterprises','Bruce Wayne','b.wayne@wayne.com','$18,500,000','Proposal','high','2026-06-30','Sarah Blackwell','North East','Enterprise Suite','Gotham infrastructure'],
]
for r in pipeline:
    ws.append(r)
for col, w in zip(range(1, 13), [8, 22, 20, 28, 14, 14, 14, 18, 18, 14, 18, 30]):
    ws.column_dimensions[get_column_letter(col)].width = w
wb3.save('sampleFile/sales_pipeline.xlsx')
print('OK sales_pipeline.xlsx')
print('All 3 test files created successfully!')
