"""
Phase 5 Multi-File Test Dataset Generator
==========================================
Creates a realistic "Invoice Processing — Q1 2025" test job with 5 files:

  1. invoices_jan_2025.csv      — Clean CSV, baseline schema (12 invoices)
  2. invoices_feb_2025.xlsx     — Excel with DIFFERENT column names for same fields
                                   (tests column aliasing / schema merge)
  3. invoices_mar_2025.csv      — CSV with an EXTRA column (discount_pct) not in
                                   Jan/Feb (tests schema evolution — new field added)
  4. vendor_payments_q1_2025.csv — Payment records linked to invoices, RELATED but
                                   distinct schema (tests multi-schema job)
  5. invoice_summary_q1.pdf      — PDF with a summary table (tests PDF/RAG pipeline)

Schema coverage by file:
  invoice_id     : jan ✓  feb ✓ (as inv_number)   mar ✓   payments ✓ (as invoice_ref)
  client_name    : jan ✓  feb ✓ (as customer)      mar ✓   payments ✗
  issue_date     : jan ✓  feb ✓                    mar ✓   payments ✗
  due_date       : jan ✓  feb ✓                    mar ✓   payments ✗
  amount         : jan ✓  feb ✓ (as net_amount)    mar ✓   payments ✓ (as amount_paid)
  tax_rate       : jan ✓  feb ✓                    mar ✓   payments ✗
  total          : jan ✓  feb ✓ (as gross_total)   mar ✓   payments ✗
  status         : jan ✓  feb ✓                    mar ✓   payments ✗
  payment_method : jan ✗  feb ✗                    mar ✗   payments ✓
  payment_date   : jan ✗  feb ✗                    mar ✗   payments ✓
  discount_pct   : jan ✗  feb ✗                    mar ✓   payments ✗  ← new in mar
"""

import csv
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = "sampleFile/test_job_invoice_q1_2025"
os.makedirs(OUT, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# FILE 1: invoices_jan_2025.csv  — clean, canonical schema
# ─────────────────────────────────────────────────────────────────────────────
jan_rows = [
    ["invoice_id", "client_name", "issue_date", "due_date", "amount", "tax_rate", "total", "status", "notes"],
    ["INV-2501", "Northwind Traders",  "2025-01-03", "2025-01-18", "5200.00", "10%", "5720.00",  "Paid",    "Early payment discount"],
    ["INV-2502", "Contoso Ltd",        "2025-01-05", "2025-01-20", "3400.00", "10%", "3740.00",  "Paid",    ""],
    ["INV-2503", "Fabrikam Inc",       "2025-01-08", "2025-01-23", "8750.00", "10%", "9625.00",  "Unpaid",  "Net 15 terms"],
    ["INV-2504", "Adventure Works",    "2025-01-10", "2025-01-25", "12000.00","10%", "13200.00", "Paid",    "Recurring monthly"],
    ["INV-2505", "Tailspin Toys",      "2025-01-12", "2025-01-27", "2150.00", "10%", "2365.00",  "Pending", "Awaiting PO"],
    ["INV-2506", "Wide World Imports", "2025-01-15", "2025-01-30", "6600.00", "10%", "7260.00",  "Unpaid",  ""],
    ["INV-2507", "Southridge Video",   "2025-01-17", "2025-02-01", "4400.00", "10%", "4840.00",  "Pending", "Check with AM"],
    ["INV-2508", "Blue Yonder Air",    "2025-01-20", "2025-02-04", "18500.00","10%", "20350.00", "Paid",    "Wire transfer confirmed"],
    ["INV-2509", "Graphic Design Inc", "2025-01-22", "2025-02-06", "1200.00", "10%", "1320.00",  "Unpaid",  ""],
    ["INV-2510", "Lucerne Publishing", "2025-01-24", "2025-02-08", "9750.00", "10%", "10725.00", "Paid",    "Invoice resent x2"],
    ["INV-2511", "Coho Winery",        "2025-01-27", "2025-02-11", "3800.00", "10%", "4180.00",  "Paid",    ""],
    ["INV-2512", "Humongous Insurance","2025-01-30", "2025-02-14", "22000.00","10%", "24200.00", "Pending", "Legal review"],
]
with open(f"{OUT}/invoices_jan_2025.csv", "w", newline="", encoding="utf-8") as f:
    csv.writer(f).writerows(jan_rows)
print("OK invoices_jan_2025.csv")


# ─────────────────────────────────────────────────────────────────────────────
# FILE 2: invoices_feb_2025.xlsx  — Excel, DIFFERENT column names (alias test)
#   inv_number  ≈ invoice_id
#   customer    ≈ client_name
#   net_amount  ≈ amount
#   gross_total ≈ total
#   Also: amounts stored as numbers (no $), dates as DD/MM/YYYY strings
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Feb Invoices"

headers_feb = ["inv_number", "customer", "issue_date", "due_date",
               "net_amount", "tax_rate", "gross_total", "status", "region"]
ws.append(headers_feb)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="2E75B6")

feb_rows = [
    ["INV-2513", "City Power & Light",  "03/02/2025", "18/02/2025",  7100.00, 0.10,  7810.00, "Unpaid",  "North"],
    ["INV-2514", "Datum Corporation",   "05/02/2025", "20/02/2025",  5500.00, 0.10,  6050.00, "Paid",    "Central"],
    ["INV-2515", "Margie Travel",       "07/02/2025", "22/02/2025",  1875.00, 0.10,  2062.50, "Paid",    "South"],
    ["INV-2516", "Proseware Inc",       "10/02/2025", "25/02/2025", 14300.00, 0.10, 15730.00, "Unpaid",  "North"],
    ["INV-2517", "School of Fine Art",  "12/02/2025", "27/02/2025",  4200.00, 0.10,  4620.00, "Pending", "East"],
    ["INV-2518", "Wingtip Toys",        "14/02/2025", "01/03/2025",  6600.00, 0.10,  7260.00, "Unpaid",  "West"],
    ["INV-2519", "Trey Research",       "16/02/2025", "03/03/2025",  3150.00, 0.10,  3465.00, "Paid",    "Central"],
    ["INV-2520", "Bellows College",     "18/02/2025", "05/03/2025", 11000.00, 0.10, 12100.00, "Pending", "North"],
    ["INV-2521", "A Datum Corp",        "20/02/2025", "07/03/2025",  2700.00, 0.10,  2970.00, "Paid",    "South"],
    ["INV-2522", "Consolidated Msg",    "22/02/2025", "09/03/2025",  8900.00, 0.10,  9790.00, "Paid",    "East"],
    ["INV-2523", "Fourth Coffee",       "24/02/2025", "11/03/2025", 16500.00, 0.10, 18150.00, "Unpaid",  "West"],
    ["INV-2524", "Relecloud",           "26/02/2025", "13/03/2025",  5250.00, 0.10,  5775.00, "Paid",    "Central"],
]
for row in feb_rows:
    ws.append(row)
for col_idx, width in enumerate([12, 24, 14, 14, 13, 10, 13, 10, 10], start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width
wb.save(f"{OUT}/invoices_feb_2025.xlsx")
print("OK invoices_feb_2025.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# FILE 3: invoices_mar_2025.csv  — CSV, adds NEW column: discount_pct (schema evolution)
#   Also uses mixed date formats (MM/DD/YYYY for some, YYYY-MM-DD for others)
#   to stress-test the multi-pass date parser
# ─────────────────────────────────────────────────────────────────────────────
mar_rows = [
    ["invoice_id", "client_name", "issue_date", "due_date", "amount", "tax_rate", "total", "discount_pct", "status", "notes"],
    ["INV-2525", "Northwind Traders",  "03/01/2025", "03/16/2025", "5400.00",  "10%", "5940.00",  "0%",   "Unpaid",  "Follow-up sent 3/1"],
    ["INV-2526", "Contoso Ltd",        "2025-03-03", "2025-03-18", "3600.00",  "10%", "3960.00",  "5%",   "Paid",    "Loyalty discount"],
    ["INV-2527", "Fabrikam Inc",       "03/05/2025", "03/20/2025", "9100.00",  "10%", "9555.00",  "5%",   "Paid",    "Volume discount"],
    ["INV-2528", "Adventure Works",    "2025-03-07", "2025-03-22", "12500.00", "10%", "13750.00", "0%",   "Unpaid",  ""],
    ["INV-2529", "Tailspin Toys",      "03/09/2025", "03/24/2025", "2300.00",  "10%", "2530.00",  "0%",   "Pending", "New PO received"],
    ["INV-2530", "Wide World Imports", "2025-03-11", "2025-03-26", "7200.00",  "10%", "7920.00",  "10%",  "Paid",    "Preferred vendor"],
    ["INV-2531", "Southridge Video",   "03/13/2025", "03/28/2025", "4600.00",  "10%", "5060.00",  "0%",   "Paid",    ""],
    ["INV-2532", "Blue Yonder Air",    "2025-03-15", "2025-03-30", "19000.00", "10%", "20900.00", "0%",   "Unpaid",  "Dispute open"],
    ["INV-2533", "Graphic Design Inc", "03/17/2025", "04/01/2025", "1350.00",  "10%", "1485.00",  "0%",   "Paid",    ""],
    ["INV-2534", "Lucerne Publishing", "2025-03-19", "2025-04-03", "10200.00", "10%", "11220.00", "5%",   "Unpaid",  ""],
    ["INV-2535", "Coho Winery",        "03/21/2025", "04/05/2025", "4000.00",  "10%", "4400.00",  "0%",   "Paid",    ""],
    ["INV-2536", "Humongous Insurance","2025-03-24", "2025-04-08", "23000.00", "10%", "25300.00", "0%",   "Pending", "Board approval needed"],
]
with open(f"{OUT}/invoices_mar_2025.csv", "w", newline="", encoding="utf-8") as f:
    csv.writer(f).writerows(mar_rows)
print("OK invoices_mar_2025.csv")


# ─────────────────────────────────────────────────────────────────────────────
# FILE 4: vendor_payments_q1_2025.csv  — Payment records, related but DISTINCT schema
#   invoice_ref ≈ invoice_id  (links to invoice files)
#   amount_paid ≈ amount/total
#   NEW fields: payment_date, payment_method, bank_ref, days_to_pay
# ─────────────────────────────────────────────────────────────────────────────
payments_rows = [
    ["payment_id", "invoice_ref", "vendor_name", "payment_date", "amount_paid", "payment_method", "bank_ref", "days_to_pay", "notes"],
    ["PAY-001", "INV-2501", "Northwind Traders",  "2025-01-16", "5720.00",  "Wire Transfer", "BNK-WT-44821", 13, "Early — 2 days before due"],
    ["PAY-002", "INV-2502", "Contoso Ltd",        "2025-01-19", "3740.00",  "Direct Debit",  "BNK-DD-44824", 14, ""],
    ["PAY-003", "INV-2504", "Adventure Works",    "2025-01-23", "13200.00", "Wire Transfer", "BNK-WT-44828", 13, "Recurring auto-pay"],
    ["PAY-004", "INV-2508", "Blue Yonder Air",    "2025-02-03", "20350.00", "Wire Transfer", "BNK-WT-44849", 14, "Confirmed by AP team"],
    ["PAY-005", "INV-2510", "Lucerne Publishing", "2025-02-07", "10725.00", "Cheque",        "CHQ-00291",    14, "Cheque posted 2/5"],
    ["PAY-006", "INV-2511", "Coho Winery",        "2025-02-10", "4180.00",  "Direct Debit",  "BNK-DD-44856", 14, ""],
    ["PAY-007", "INV-2513", "City Power & Light",  None,        None,        None,            None,          None,"Overdue — escalated"],
    ["PAY-008", "INV-2514", "Datum Corporation",  "2025-02-19", "6050.00",  "Direct Debit",  "BNK-DD-44874", 14, ""],
    ["PAY-009", "INV-2515", "Margie Travel",      "2025-02-20", "2062.50",  "Wire Transfer", "BNK-WT-44875", 13, ""],
    ["PAY-010", "INV-2519", "Trey Research",      "2025-03-02", "3465.00",  "Direct Debit",  "BNK-DD-44888", 14, "Overpaid — credit note raised"],
    ["PAY-011", "INV-2521", "A Datum Corp",       "2025-03-06", "2970.00",  "Wire Transfer", "BNK-WT-44894", 14, ""],
    ["PAY-012", "INV-2522", "Consolidated Msg",   "2025-03-08", "9790.00",  "Direct Debit",  "BNK-DD-44896", 14, "Direct debit confirmed"],
    ["PAY-013", "INV-2524", "Relecloud",          "2025-03-12", "5775.00",  "Wire Transfer", "BNK-WT-44900", 14, ""],
    ["PAY-014", "INV-2526", "Contoso Ltd",        "2025-03-17", "3762.00",  "Direct Debit",  "BNK-DD-44905", 14, "Loyalty discount applied"],
    ["PAY-015", "INV-2527", "Fabrikam Inc",       "2025-03-19", "9079.25",  "Wire Transfer", "BNK-WT-44907", 14, "Volume discount applied"],
    ["PAY-016", "INV-2530", "Wide World Imports", "2025-03-25", "7128.00",  "Direct Debit",  "BNK-DD-44913", 14, "Preferred vendor rate"],
    ["PAY-017", "INV-2531", "Southridge Video",   "2025-03-27", "5060.00",  "Wire Transfer", "BNK-WT-44915", 14, ""],
]
with open(f"{OUT}/vendor_payments_q1_2025.csv", "w", newline="", encoding="utf-8") as f:
    csv.writer(f).writerows(payments_rows)
print("OK vendor_payments_q1_2025.csv")


# ─────────────────────────────────────────────────────────────────────────────
# FILE 5: invoice_summary_q1.pdf  — PDF with summary table (tests PDF/RAG pipeline)
# ─────────────────────────────────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        f"{OUT}/invoice_summary_q1.pdf",
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                  fontSize=16, alignment=TA_CENTER, spaceAfter=6)
    sub_style   = ParagraphStyle("Sub", parent=styles["Normal"],
                                  fontSize=10, alignment=TA_CENTER, spaceAfter=16, textColor=colors.grey)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"],
                                    fontSize=12, spaceBefore=16, spaceAfter=6)
    body_style  = ParagraphStyle("Body", parent=styles["Normal"], fontSize=9, spaceAfter=8)

    elements = []
    elements.append(Paragraph("Acme Corp — Invoice Processing Report", title_style))
    elements.append(Paragraph("Q1 2025 (January – March) | Prepared: April 1, 2025", sub_style))

    elements.append(Paragraph("1. Executive Summary", section_style))
    elements.append(Paragraph(
        "This report summarises all invoices issued and payments received during Q1 2025. "
        "A total of 36 invoices were processed across three months, with a combined gross value "
        "of £518,435.00. Payment compliance stood at 72% with 4 invoices remaining overdue "
        "beyond their due date. The average days-to-pay across settled invoices was 13.6 days.",
        body_style
    ))

    elements.append(Paragraph("2. Monthly Breakdown", section_style))
    monthly_data = [
        ["Month",    "Invoices Issued", "Total Net (£)", "Total Tax (£)", "Total Gross (£)", "% Paid"],
        ["January",  "12",  "98,750.00",  "9,875.00",  "108,625.00", "50%"],
        ["February", "12",  "91,075.00",  "9,107.50",  "100,182.50", "58%"],
        ["March",    "12", "100,250.00", "10,025.00",  "281,627.50", "67%"],
        ["TOTAL",    "36", "290,075.00", "29,007.50",  "318,082.50", "58%"],
    ]
    t1 = Table(monthly_data, colWidths=[3.5*cm, 3.5*cm, 4*cm, 3.5*cm, 4*cm, 2.5*cm])
    t1.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("BACKGROUND",   (0,-1), (-1,-1), colors.HexColor("#DEEAF1")),
        ("FONTNAME",     (0,-1), (-1,-1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(t1)

    elements.append(Paragraph("3. Outstanding Invoices (Overdue)", section_style))
    overdue_data = [
        ["Invoice ID",  "Client",             "Due Date",   "Amount (£)", "Days Overdue", "Action"],
        ["INV-2503",    "Fabrikam Inc",        "2025-01-23", "9,625.00",   "68",           "Escalated to collections"],
        ["INV-2506",    "Wide World Imports",  "2025-01-30", "7,260.00",   "61",           "Second notice sent"],
        ["INV-2513",    "City Power & Light",  "2025-02-18", "7,810.00",   "41",           "Dispute in progress"],
        ["INV-2516",    "Proseware Inc",       "2025-02-25", "15,730.00",  "34",           "Awaiting PO approval"],
        ["INV-2523",    "Fourth Coffee",       "2025-03-11", "18,150.00",  "21",           "Final demand issued"],
        ["INV-2528",    "Adventure Works",     "2025-03-22", "13,750.00",  "10",           "Reminder sent"],
    ]
    t2 = Table(overdue_data, colWidths=[2.5*cm, 4.5*cm, 3*cm, 3*cm, 3*cm, 5*cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#C00000")),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("ALIGN",        (5,1), (5,-1), "LEFT"),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#FFF2F2")]),
    ]))
    elements.append(t2)

    elements.append(Paragraph("4. Payment Method Distribution", section_style))
    payment_data = [
        ["Payment Method", "Count", "Total Paid (£)", "% of Payments"],
        ["Wire Transfer",  "9",     "174,507.25",     "53%"],
        ["Direct Debit",   "7",     "103,050.00",     "38%"],
        ["Cheque",         "1",     "10,725.00",      "6%"],
        ["Pending/None",   "4",     "—",              "—"],
    ]
    t3 = Table(payment_data, colWidths=[4.5*cm, 2.5*cm, 4.5*cm, 4.5*cm])
    t3.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#7030A0")),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elements.append(t3)

    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(
        "5. Notes & Observations",
        section_style
    ))
    notes_items = [
        "INV-2519 (Trey Research): Overpayment of £315 received — credit note CN-2025-001 issued on 03/03/2025.",
        "INV-2510 (Lucerne Publishing): Invoice resent twice due to bounced email. Confirmed received 01/24/2025.",
        "March invoices include a discount_pct column reflecting negotiated vendor rates introduced from March 2025 onwards.",
        "All amounts in GBP (£) unless stated otherwise. Tax rate uniformly 10% for Q1 2025.",
        "Data extracted from accounts payable system. Any discrepancies should be reported to finance@acme.com.",
    ]
    for note in notes_items:
        elements.append(Paragraph(f"• {note}", body_style))

    doc.build(elements)
    print("OK invoice_summary_q1.pdf")

except Exception as e:
    print(f"WARN: PDF generation failed ({e}) — skipping invoice_summary_q1.pdf")


# ─────────────────────────────────────────────────────────────────────────────
# README for the test group
# ─────────────────────────────────────────────────────────────────────────────
readme = """\
# Test Job: Invoice Processing — Q1 2025

## Purpose
Multi-file test dataset for Phase 5 development.
Use job purpose: "Invoice Processing — Q1 2025"

## Files

| File | Type | Rows | Notes |
|------|------|------|-------|
| invoices_jan_2025.csv | CSV | 12 | Clean canonical schema — baseline |
| invoices_feb_2025.xlsx | Excel | 12 | DIFFERENT column names (alias test): inv_number, customer, net_amount, gross_total |
| invoices_mar_2025.csv | CSV | 12 | Adds NEW column: discount_pct (schema evolution test). Mixed date formats. |
| vendor_payments_q1_2025.csv | CSV | 17 | Related but distinct schema: payment_date, payment_method, bank_ref, days_to_pay |
| invoice_summary_q1.pdf | PDF | 4 tables | PDF with summary tables (PDF/RAG pipeline test) |

## Schema Coverage

| Column | Jan | Feb | Mar | Payments |
|--------|-----|-----|-----|----------|
| invoice_id | ✓ | ✓ (as inv_number) | ✓ | ✓ (as invoice_ref) |
| client_name | ✓ | ✓ (as customer) | ✓ | - |
| issue_date | ✓ | ✓ | ✓ | - |
| due_date | ✓ | ✓ | ✓ | - |
| amount | ✓ | ✓ (as net_amount) | ✓ | ✓ (as amount_paid) |
| tax_rate | ✓ | ✓ | ✓ | - |
| total | ✓ | ✓ (as gross_total) | ✓ | - |
| status | ✓ | ✓ | ✓ | - |
| discount_pct | - | - | ✓ NEW | - |
| payment_date | - | - | - | ✓ NEW |
| payment_method | - | - | - | ✓ NEW |
| bank_ref | - | - | - | ✓ NEW |
| days_to_pay | - | - | - | ✓ NEW |
| region | - | ✓ NEW | - | - |

## What Each File Tests
- **Jan CSV**: Baseline — confirms standard pipeline still works
- **Feb Excel**: Column aliasing — schema merge must detect inv_number ≈ invoice_id, customer ≈ client_name, etc.
- **Mar CSV**: Schema evolution — discount_pct should be added to job schema as a new nullable field after Mar is processed
- **Payments CSV**: Multi-schema — linked via invoice_ref but represents a different entity (payments, not invoices)
- **Summary PDF**: PDF/RAG pipeline — contains 4 tables, tests chunking, embedding, and retrieval
"""
with open(f"{OUT}/README.md", "w", encoding="utf-8") as f:
    f.write(readme)

print("\nAll Phase 5 test files created in:", OUT)
print("Folder contents:")
for name in sorted(os.listdir(OUT)):
    path = os.path.join(OUT, name)
    size_kb = os.path.getsize(path) / 1024
    print(f"  {name:<40} {size_kb:>7.1f} KB")
